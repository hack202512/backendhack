from typing import List
from io import BytesIO
from datetime import datetime, time
import uuid
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from models.models import RegistryCounter, CountyOffice
from context.db import get_db
from functions.auth import get_current_user_token
from models.models import FoundItem, User
from schemas.found_item_form import FoundItemFormRequest, FoundItemFormResponse

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    openpyxl = None


router = APIRouter(prefix="/found-item-forms", tags=["found-item-forms"])


def _naive(dt):
    if dt and getattr(dt, "tzinfo", None):
        return dt.replace(tzinfo=None)
    return dt

def next_registry_number(db: Session, office: CountyOffice, dt: datetime | None = None) -> str:
    dt = dt or datetime.utcnow()
    year = dt.year

    counter = (
        db.query(RegistryCounter)
        .filter_by(county_office_id=office.id, year=year)
        .with_for_update()
        .first()
    )

    if not counter:
        counter = RegistryCounter(county_office_id=office.id, year=year, value=0)
        db.add(counter)
        db.flush()

    counter.value += 1
    seq = counter.value

    yy = str(year)[-2:]
    code = (office.code or "XX").upper()

    return f"RZ{yy}{code}{seq:04d}"

def require_user(
    token_data: dict = Depends(get_current_user_token),
    db: Session = Depends(get_db),
) -> User:
    user_id = token_data.get("user_id")
    if not user_id:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token payload")

    user = db.query(User).filter(User.id == int(user_id)).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
    return user


def to_form_response(i: FoundItem) -> FoundItemFormResponse:
    created_at = getattr(i, "created_at", None)
    if not created_at:
        from datetime import timezone
        created_at = datetime.now(timezone.utc)

    found_date = getattr(i, "found_date", None)

    return FoundItemFormResponse(
        id=str(getattr(i, "id")),
        registry_number=getattr(i, "registry_number", None),
        item_name=getattr(i, "item_name", "") or getattr(i, "name", "") or "",
        item_color=getattr(i, "item_color", None),
        item_brand=getattr(i, "item_brand", None),
        found_location=getattr(i, "found_location", None),
        found_date=found_date,
        found_time=getattr(i, "found_time", None),
        circumstances=getattr(i, "circumstances", None),
        found_by_firstname=getattr(i, "found_by_firstname", None),
        found_by_lastname=getattr(i, "found_by_lastname", None),
        found_by_phonenumber=getattr(i, "found_by_phonenumber", None),
        created_at=created_at,
    )


def _fmt_found(dt):
    if not dt:
        return ""
    try:
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(dt)

@router.post("/", response_model=FoundItemFormResponse, status_code=201)
def add_found_item(
    payload: FoundItemFormRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    if not hasattr(FoundItem, "county_office_id") or not hasattr(FoundItem, "registry_number"):
        raise HTTPException(500, detail="Model FoundItem missing county_office_id/registry_number fields")
    if not hasattr(User, "county_offices"):
        raise HTTPException(500, detail="Model User missing county_offices relation")
    if not hasattr(CountyOffice, "code"):
        raise HTTPException(500, detail="Model CountyOffice missing code field")

    office = current_user.county_offices[0] if current_user.county_offices else None
    if not office:
        raise HTTPException(400, detail="User has no county office assigned")

    item = FoundItem()

    item.item_name = payload.item_name.strip()
    item.item_color = payload.item_color.strip() if payload.item_color else None
    item.item_brand = payload.item_brand.strip() if payload.item_brand else None
    item.found_location = payload.found_location.strip() if payload.found_location else None

    if payload.found_time:
        try:
            time_obj = datetime.strptime(payload.found_time, "%H:%M").time()
            item.found_date = datetime.combine(payload.found_date, time_obj)
        except ValueError:
            item.found_date = datetime.combine(payload.found_date, time.min)
    else:
        item.found_date = datetime.combine(payload.found_date, time.min)

    item.found_time = payload.found_time.strip() if payload.found_time else None
    item.circumstances = payload.circumstances.strip() if payload.circumstances else None
    item.found_by_firstname = payload.found_by_firstname.strip() if payload.found_by_firstname else None
    item.found_by_lastname = payload.found_by_lastname.strip() if payload.found_by_lastname else None
    item.found_by_phonenumber = payload.found_by_phonenumber.strip() if payload.found_by_phonenumber else None

    item.user_id = current_user.id

    item.county_office_id = office.id

    try:
        item.registry_number = next_registry_number(db, office)
    except Exception as e:
        db.rollback()
        raise HTTPException(500, detail=f"Failed to generate registry number: {e}")

    db.add(item)
    db.commit()
    db.refresh(item)

    return to_form_response(item)


@router.get("/my", response_model=List[FoundItemFormResponse])
def list_my_found_items(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    if not hasattr(FoundItem, "user_id"):
        raise HTTPException(500, detail="Model FoundItem missing user_id field")

    order_col = getattr(FoundItem, "created_at", None) or getattr(FoundItem, "id")
    items = (
        db.query(FoundItem)
        .filter(FoundItem.user_id == current_user.id)
        .order_by(order_col.desc())
        .all()
    )
    return [to_form_response(i) for i in items]


@router.get("/export")
def export_my_forms(
    format: str = Query("xlsx", pattern="^(xlsx|excel|json|csv)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    order_col = getattr(FoundItem, "created_at", None) or getattr(FoundItem, "id")
    items = (
        db.query(FoundItem)
        .filter(FoundItem.user_id == current_user.id)
        .order_by(order_col.desc())
        .all()
    )
    mapped = [to_form_response(i) for i in items]

    if format == "csv":
        import csv
        from io import StringIO

        def _fmt_dt(dt):
            if not dt:
                return ""
            try:
                return dt.replace(microsecond=0).isoformat()
            except Exception:
                return str(dt)

        output = StringIO()
        writer = csv.writer(
            output,
            delimiter=";",
            quoting=csv.QUOTE_MINIMAL,
        )

        writer.writerow([
            "ID",
            "Numer ewidencyjny",
            "Nazwa",
            "Lokalizacja",
            "Data znalezienia",
            "Utworzono",
        ])

        for m in mapped:
            writer.writerow([
                m.id,
                getattr(m, "registry_number", None) or "",
                m.item_name,
                m.found_location or "",
                _fmt_found(m.found_date),
                _fmt_dt(m.created_at),
            ])

        csv_text = output.getvalue()
        output.close()

        csv_bytes = ("\ufeff" + csv_text).encode("utf-8")

        headers = {"Content-Disposition": "attachment; filename=found_items.csv"}
        return StreamingResponse(
            BytesIO(csv_bytes),
            media_type="text/csv; charset=utf-8",
            headers=headers,
        )

    elif format == "json":
        import json

        data = []
        for m in mapped:
            data.append({
                "id": m.id,
                "registry_number": getattr(m, "registry_number", None),
                "item_name": m.item_name,
                "item_color": m.item_color,
                "item_brand": m.item_brand,
                "found_location": m.found_location,
                "found_date": m.found_date.isoformat() if m.found_date else None,
                "created_at": m.created_at.isoformat() if m.created_at else None,
            })

        buf = BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"))
        headers = {"Content-Disposition": "attachment; filename=found_items.json"}
        return StreamingResponse(
            buf,
            media_type="application/json",
            headers=headers,
        )

    if openpyxl is None:
        raise HTTPException(500, detail="openpyxl not installed. Add it to requirements.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Found items"

    headers_row = [
        "ID",
        "Registry number",
        "Item name",
        "Color",
        "Brand",
        "Found location",
        "Found date",
        "Created at",
    ]

    ws.append(headers_row)

    for m in mapped:
        ws.append([
            m.id,
            getattr(m, "registry_number", None),
            m.item_name,
            m.item_color,
            m.item_brand,
            m.found_location,
            m.found_date,
            _naive(m.created_at),
        ])


    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    table = Table(displayName="FoundItemsTable", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    header_index = {cell.value: cell.column for cell in ws[1]}
    found_date_col = header_index.get("Found date")
    created_at_col = header_index.get("Created at")

    if found_date_col:
        for row in ws.iter_rows(min_row=2, min_col=found_date_col, max_col=found_date_col):
            for cell in row:
                if cell.value:
                    cell.number_format = "yyyy-mm-dd hh:mm"

    if created_at_col:
        for row in ws.iter_rows(min_row=2, min_col=created_at_col, max_col=created_at_col):
            for cell in row:
                if cell.value:
                    cell.number_format = "yyyy-mm-dd hh:mm"

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            s = v.isoformat() if hasattr(v, "isoformat") else str(v)
            max_len = max(max_len, len(s))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 50)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    headers = {"Content-Disposition": "attachment; filename=found_items.xlsx"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/{item_id}", response_model=FoundItemFormResponse)
def get_found_item(
    item_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError:
        raise HTTPException(400, detail="Invalid item_id")

    item = (
        db.query(FoundItem)
        .filter(FoundItem.id == item_uuid, FoundItem.user_id == current_user.id)
        .first()
    )
    if not item:
        raise HTTPException(404, detail="Form not found")
    return to_form_response(item)